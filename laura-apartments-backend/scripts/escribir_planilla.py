# -*- coding: utf-8 -*-
r"""
Escribe los resultados de la homologacion (CSV que genera homologacion.ps1) en las columnas
verdes de la planilla de Fiserv. No toca el original: guarda una copia -COMPLETADA.xlsx.

Uso:
    python escribir_planilla.py
    python escribir_planilla.py --csv ruta\resultados.csv --xlsx ruta\Homologacion-Tarjetas.xlsx

Las filas 20-23 (Ley17934/IMESI) y 25-26 (EDIT PAYMENT) se marcan N/A. La 24 (ECHO TEST) ya OK.
"""
import argparse
import csv
import os
import shutil
import openpyxl

DEFAULT_XLSX = r"C:\Users\colla\Downloads\Homologación-Tarjetas.xlsx"
DEFAULT_CSV = os.path.join(os.path.dirname(__file__), "homologacion-out", "resultados.csv")

# Filas que no se ejecutan: se marcan N/A con el motivo.
NA_ROWS = {
    20: "N/A - no aplica al rubro (Ley 17934)",
    21: "N/A - no aplica al rubro (Ley 17934)",
    22: "N/A - no aplica al rubro (IMESI)",
    23: "N/A - no aplica al rubro (IMESI)",
    25: "N/A - EDIT PAYMENT no aplica a integracion hosted page",
    26: "N/A - EDIT PAYMENT no aplica a integracion hosted page",
}
ECHO_ROW = {24: ("Echo test automatico cada 5 min - OK", "Autorizada")}


def find_columns(ws):
    """Ubica la fila de encabezados y devuelve el indice de cada columna verde por nombre."""
    wanted = {
        "CodRespuesta": None, "Fecha": None, "Hora": None,
        "Referencia ecommerce": None, "AccessToken": None, "Resultado": None,
    }
    obs_idx = None  # la columna "Observaciones" de respuesta (la segunda, no la de precondiciones)
    header_row = None
    for r in range(1, 6):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if "CodRespuesta" in [str(v).strip() if v else v for v in vals]:
            header_row = r
            break
    if header_row is None:
        raise SystemExit("No encontre la fila de encabezados con 'CodRespuesta'")

    seen_codresp = False
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        v = str(v).strip() if v else v
        if v in wanted:
            wanted[v] = c
        if v == "CodRespuesta":
            seen_codresp = True
        # La columna "Observaciones" que esta despues de CodRespuesta es la de respuesta.
        if v == "Observaciones" and seen_codresp and obs_idx is None:
            obs_idx = c
    wanted["Observaciones"] = obs_idx
    return header_row, wanted


def build_row_map(xlsx_path, header_row):
    """La columna '#' son formulas (=A3+1), asi que leemos los valores CACHEADOS con
    data_only=True para mapear numero_de_transaccion -> fila de Excel."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Transacciones"]
    mapping = {}
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        try:
            mapping[int(float(v))] = r
        except (ValueError, TypeError):
            continue
    return mapping


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", default=DEFAULT_CSV)
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    args = ap.parse_args()

    base, ext = os.path.splitext(args.xlsx)
    out_path = base + "-COMPLETADA" + ext
    shutil.copyfile(args.xlsx, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb["Transacciones"]
    header_row, cols = find_columns(ws)
    row_map = build_row_map(out_path, header_row)

    def write(excel_row, colname, value):
        if value is None or value == "":
            return
        c = cols.get(colname)
        if c:
            ws.cell(row=excel_row, column=c).value = value

    # Resultados ejecutados (desde el CSV).
    results = {}
    if os.path.exists(args.csv):
        with open(args.csv, encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                results[int(row["Row"])] = row

    written = []
    for num, row in sorted(results.items()):
        er = row_map.get(num)
        if er is None:
            print(f"  OJO: no encontre la fila #{num} en la planilla")
            continue
        write(er, "Referencia ecommerce", row.get("Reference"))
        write(er, "AccessToken", row.get("Token"))
        write(er, "CodRespuesta", row.get("CodRespuesta"))
        write(er, "Resultado", row.get("Resultado"))
        write(er, "Fecha", row.get("Fecha"))
        write(er, "Hora", row.get("Hora"))
        # Observaciones: cuotas + la nota que traiga el CSV (autorizacion, anulacion, ley, etc.)
        obs = []
        if row.get("Cuotas"):
            obs.append(f"cuotas={row['Cuotas']}")
        if row.get("Obs"):
            obs.append(row["Obs"])
        write(er, "Observaciones", "; ".join(obs))
        written.append(num)

    # Filas N/A.
    for num, motivo in NA_ROWS.items():
        er = row_map.get(num)
        if er:
            write(er, "Resultado", "N/A")
            write(er, "Observaciones", motivo)
            written.append(num)

    # Echo test.
    for num, (obs, res) in ECHO_ROW.items():
        er = row_map.get(num)
        if er:
            write(er, "Resultado", res)
            write(er, "Observaciones", obs)
            written.append(num)

    wb.save(out_path)
    print(f"Listo. Escritas {len(written)} filas: {sorted(written)}")
    print(f"Guardado en: {out_path}")


if __name__ == "__main__":
    main()
